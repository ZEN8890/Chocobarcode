import pandas as pd
from barcode.ean import EAN13
from barcode.writer import ImageWriter
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from io import BytesIO
import os
import random
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk
import threading
import time

# --- Konfigurasi Awal (Bisa disesuaikan di GUI nanti) ---
DEFAULT_INPUT_FILE = 'produk_barcode_lengkap.xlsx'
OUTPUT_FILE_NAME = 'barcode_list_chocobarcode.xlsx' 
BARCODE_COLUMN_NAME = 'Barcode (EAN-13)'
PRODUCT_NAME_COLUMN_NAME = 'Nama Produk'
OUTPUT_BARCODE_IMAGE_COLUMN_HEADER = 'Gambar Barcode'

IMAGE_WIDTH_PIXELS = 250
IMAGE_HEIGHT_PIXELS = 180

# --- Fungsi Logika Barcode ---

def calculate_ean13_checksum(first12digits):
    """
    Menghitung digit checksum untuk 12 digit pertama barcode EAN-13.
    """
    sum_odd = 0
    sum_even = 0
    for i, digit in enumerate(first12digits):
        if (i + 1) % 2 == 1:
            sum_odd += int(digit)
        else:
            sum_even += int(digit)
    total = sum_odd * 1 + sum_even * 3
    checksum = (10 - (total % 10)) % 10
    return checksum

def generate_valid_ean13_string(barcode_number):
    """
    Memvalidasi dan mengoreksi barcode agar menjadi 13 digit EAN-13 yang valid secara checksum.
    Fungsi ini akan mengembalikan string 13 digit atau None jika input tidak valid.
    """
    barcode_str = str(barcode_number).strip()
    
    # Periksa apakah input hanya terdiri dari digit
    if not barcode_str.isdigit():
        return None # Mengembalikan None jika bukan angka

    if len(barcode_str) == 13:
        first12 = barcode_str[:-1]
        input_checksum = int(barcode_str[-1])
        correct_checksum = calculate_ean13_checksum(first12)
        if input_checksum == correct_checksum:
            return barcode_str
        else:
            # Jika checksum tidak cocok, koreksi barcode dengan checksum yang benar
            corrected_barcode = first12 + str(correct_checksum)
            return corrected_barcode
    
    elif len(barcode_str) == 12:
        # Jika 12 digit, hitung dan tambahkan checksum
        checksum = calculate_ean13_checksum(barcode_str)
        valid_barcode = barcode_str + str(checksum)
        return valid_barcode
    
    else: 
        # Untuk kasus lain (kurang dari 12 atau lebih dari 13)
        return None 

def generate_new_unique_ean13(existing_barcodes_set, log_callback=None):
    """
    Menghasilkan barcode EAN-13 baru yang unik dan valid,
    memastikan tidak ada di dalam set existing_barcodes_set.
    """
    while True:
        # Hasilkan 12 digit angka acak. EAN-13 dimulai dengan 12 digit + 1 digit checksum.
        random_12_digits = str(random.randint(10**11, 10**12 - 1))
        
        # Hitung checksum untuk 12 digit acak ini
        checksum = calculate_ean13_checksum(random_12_digits)
        new_ean13 = random_12_digits + str(checksum)
        
        # Periksa apakah barcode yang baru dihasilkan ini unik
        if new_ean13 not in existing_barcodes_set:
            if log_callback:
                log_callback(f"    Info: Barcode baru yang unik dihasilkan: '{new_ean13}'.")
            return new_ean13
        else:
            if log_callback:
                log_callback(f"    Peringatan: Barcode acak '{new_ean13}' sudah ada, mencoba lagi...")


def generate_ean13_image_buffer(barcode_number_str):
    """
    Membuat gambar barcode dan mengembalikannya sebagai buffer memori.
    """
    ean = EAN13(barcode_number_str, writer=ImageWriter()) 
    buffer = BytesIO()
    ean.write(buffer, {
        'module_width': 0.25,
        'module_height': 12,
        'quiet_zone': 6,
        'text_distance': 5.0,
        'font_size': 12,
        'background': 'white',
        'foreground': 'black',
        'write_text': True,
        'text': barcode_number_str
    })
    buffer.seek(0)
    return buffer

# --- Kelas Aplikasi GUI ---

class BarcodeApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Chocobarcode EAN-13")
        self.root.geometry("800x600")

        self.main_frame = ttk.Frame(root, padding="10")
        self.main_frame.pack(fill=tk.BOTH, expand=True)

        self.file_frame = ttk.LabelFrame(self.main_frame, text="Pengaturan File", padding="10")
        self.file_frame.pack(fill=tk.X, pady=10)

        ttk.Label(self.file_frame, text="File Input Excel:").grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
        self.input_file_entry = ttk.Entry(self.file_frame, width=50)
        self.input_file_entry.grid(row=0, column=1, padx=5, pady=5, sticky=(tk.W, tk.E))
        self.input_file_entry.insert(0, DEFAULT_INPUT_FILE)
        ttk.Button(self.file_frame, text="Browse Input", command=self.browse_input_file).grid(row=0, column=2, padx=5, pady=5)

        ttk.Label(self.file_frame, text="Folder Output:").grid(row=1, column=0, padx=5, pady=5, sticky=tk.W)
        self.output_folder_entry = ttk.Entry(self.file_frame, width=50)
        self.output_folder_entry.grid(row=1, column=1, padx=5, pady=5, sticky=(tk.W, tk.E))
        self.output_folder_entry.insert(0, os.getcwd()) 
        ttk.Button(self.file_frame, text="Browse Output Folder", command=self.browse_output_folder).grid(row=1, column=2, padx=5, pady=5)
        
        self.export_format_button = ttk.Button(self.file_frame, text="Export Format Kosong", command=self.export_empty_format_thread)
        self.export_format_button.grid(row=2, column=0, columnspan=3, pady=10)

        self.file_frame.grid_columnconfigure(1, weight=1)

        self.process_frame = ttk.LabelFrame(self.main_frame, text="Kontrol Proses", padding="10")
        self.process_frame.pack(fill=tk.X, pady=10)

        self.start_button = ttk.Button(self.process_frame, text="Mulai Generate Barcode", command=self.start_generation_thread)
        self.start_button.pack(pady=10)

        self.progress_bar = ttk.Progressbar(self.process_frame, orient="horizontal", mode="determinate")
        self.progress_bar.pack(fill=tk.X, pady=5)

        self.log_frame = ttk.LabelFrame(self.main_frame, text="Log Proses", padding="10")
        self.log_frame.pack(fill=tk.BOTH, expand=True, pady=10)

        self.log_text = scrolledtext.ScrolledText(self.log_frame, wrap=tk.WORD, width=80, height=15, state='disabled')
        self.log_text.pack(fill=tk.BOTH, expand=True)

        self.status_frame = ttk.LabelFrame(self.main_frame, text="Status & Hasil", padding="10")
        self.status_frame.pack(fill=tk.X, pady=10)

        self.status_label = ttk.Label(self.status_frame, text="Status: Siap", font=('Helvetica', 10, 'bold'))
        self.status_label.pack(pady=5)

        self.results_label = ttk.Label(self.status_frame, text="")
        self.results_label.pack(pady=5)

        self.credit_label = ttk.Label(self.main_frame, text="Credit by Steven Gunawan", foreground="gray", font=('Helvetica', 9, 'italic'))
        self.credit_label.pack(side=tk.BOTTOM, pady=5)

    def log_message(self, message):
        self.log_text.config(state='normal')
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
        self.log_text.config(state='disabled')
        self.root.update_idletasks()

    def browse_input_file(self):
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if file_path:
            self.input_file_entry.delete(0, tk.END)
            self.input_file_entry.insert(0, file_path)
            self.log_message(f"File input dipilih: {file_path}")

    def browse_output_folder(self):
        folder_path = filedialog.askdirectory()
        if folder_path:
            self.output_folder_entry.delete(0, tk.END)
            self.output_folder_entry.insert(0, folder_path)
            self.log_message(f"Folder output dipilih: {folder_path}")

    def _set_gui_processing_state(self, is_processing):
        state = 'disabled' if is_processing else 'normal'
        self.start_button.config(state=state)
        self.export_format_button.config(state=state)
        if is_processing:
            self.log_text.config(state='normal')
            self.log_text.delete('1.0', tk.END)
            self.log_text.config(state='disabled')
            self.status_label.config(text="Status: Memulai...")
            self.results_label.config(text="")
            self.progress_bar['value'] = 0
            self.progress_bar['mode'] = 'indeterminate'
            self.progress_bar.start()
        else:
            self.progress_bar.stop()
            self.progress_bar.config(value=0)
            self.status_label.config(text="Status: Siap")

    def start_generation_thread(self):
        input_file = self.input_file_entry.get()
        output_folder = self.output_folder_entry.get()
        
        output_file = os.path.join(output_folder, OUTPUT_FILE_NAME)

        if not input_file or not output_folder:
            messagebox.showerror("Error", "Harap pilih file input dan folder output.")
            return

        if not os.path.exists(input_file):
            messagebox.showerror("Error", f"File input '{input_file}' tidak ditemukan.")
            return
        
        if not os.path.isdir(output_folder):
            try:
                os.makedirs(output_folder)
                self.log_message(f"Folder output '{output_folder}' berhasil dibuat.")
            except Exception as e:
                messagebox.showerror("Error", f"Gagal membuat folder output '{output_folder}': {e}")
                return

        self._set_gui_processing_state(True)
        threading.Thread(target=self._generate_barcodes_process, args=(input_file, output_file)).start()

    def _generate_barcodes_process(self, input_file, output_file):
        self.log_message("Memulai proses pembuatan barcode Excel...")
        self.log_message("PENTING: Barcode yang valid (13 digit, checksum benar) dari file input akan dipertahankan.")
        self.log_message(f"File output akan disimpan sebagai: {output_file}")

        try:
            df = pd.read_excel(input_file, dtype={BARCODE_COLUMN_NAME: str})
            total_rows = len(df)
            self.log_message(f"Berhasil membaca {total_rows} baris dari '{input_file}'.")

            self.root.after(100, lambda: self.progress_bar.stop())
            self.root.after(100, lambda: self.progress_bar.config(mode='determinate', maximum=total_rows))

            wb = Workbook()
            ws = wb.active
            ws.title = "Produk Barcode"

            headers = [PRODUCT_NAME_COLUMN_NAME, BARCODE_COLUMN_NAME, OUTPUT_BARCODE_IMAGE_COLUMN_HEADER]
            ws.append(headers)

            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = IMAGE_WIDTH_PIXELS / 7 

            # Mengumpulkan barcode yang sudah ada untuk memeriksa duplikasi.
            # Menggunakan set untuk pencarian yang efisien.
            processed_barcodes = set()
            generated_new_barcode_count = 0
            failed_processing_count = 0
            successful_processing_count = 0
            
            # --- START: Perbaikan logika pemrosesan barcode ---
            
            # Pra-proses semua barcode dari input untuk mengidentifikasi duplikat awal
            existing_barcodes = df[BARCODE_COLUMN_NAME].astype(str).tolist()
            
            for index, row in df.iterrows():
                product_name = str(row[PRODUCT_NAME_COLUMN_NAME]).strip()
                original_barcode = str(row[BARCODE_COLUMN_NAME]).strip()
                
                self.log_message(f"\nMemproses: '{product_name}' - Barcode Asli: '{original_barcode}'")

                final_barcode_for_excel = None
                try:
                    # Coba validasi barcode asli dari input Excel
                    validated_barcode = generate_valid_ean13_string(original_barcode)

                    if validated_barcode: # Barcode asli valid atau sudah dikoreksi checksum
                        if validated_barcode in processed_barcodes:
                            self.log_message(f"    Peringatan: Barcode EAN-13 yang divalidasi '{validated_barcode}' adalah duplikat. Menghasilkan barcode baru.")
                            final_barcode_for_excel = generate_new_unique_ean13(processed_barcodes, self.log_message)
                            generated_new_barcode_count += 1
                        else:
                            # Jika valid dan belum diproses, gunakan barcode ini
                            final_barcode_for_excel = validated_barcode
                            self.log_message(f"    Info: Barcode awal divalidasi dan digunakan: '{final_barcode_for_excel}'.")
                    else:
                        # Jika barcode asli tidak valid
                        self.log_message(f"    Peringatan: Barcode asli '{original_barcode}' tidak valid atau terlalu pendek/panjang. Menghasilkan barcode baru.")
                        final_barcode_for_excel = generate_new_unique_ean13(processed_barcodes, self.log_message)
                        generated_new_barcode_count += 1

                    if final_barcode_for_excel:
                        processed_barcodes.add(final_barcode_for_excel) # Tambahkan ke set barcode yang sudah diproses

                        buffer = generate_ean13_image_buffer(final_barcode_for_excel)
                        
                        current_row_in_excel = ws.max_row + 1
                        ws.append([product_name, final_barcode_for_excel, ''])

                        img = OpenpyxlImage(buffer)
                        img.width = IMAGE_WIDTH_PIXELS
                        img.height = IMAGE_HEIGHT_PIXELS

                        cell_ref = f'C{current_row_in_excel}'
                        ws.add_image(img, cell_ref)
                        
                        ws.row_dimensions[current_row_in_excel].height = IMAGE_HEIGHT_PIXELS * 0.7

                        self.log_message(f"    Berhasil dibuat: Barcode EAN-13 '{final_barcode_for_excel}' dan gambar disisipkan di baris {current_row_in_excel}.")
                        successful_processing_count += 1
                    else:
                        self.log_message(f"    Gagal memproses barcode '{original_barcode}': Barcode final tidak dapat ditentukan.")
                        ws.append([product_name, original_barcode, "GAGAL GENERATE BARCODE"])
                        failed_processing_count += 1

                except Exception as e:
                    self.log_message(f"    Gagal memproses barcode '{original_barcode}': {str(e)}")
                    ws.append([product_name, original_barcode, "GAGAL GENERATE BARCODE"])
                    failed_processing_count += 1
                
                self.root.after(0, lambda idx=index: self.progress_bar.config(value=idx + 1))
            
            # --- END: Perbaikan logika pemrosesan barcode ---

            wb.save(output_file)
            
            self.log_message(f"\nProses selesai! File Excel '{output_file}' telah berhasil dibuat.")
            self.log_message(f"Jumlah produk yang berhasil diproses: {successful_processing_count}")
            self.log_message(f"Jumlah barcode baru yang dihasilkan: {generated_new_barcode_count}")
            self.log_message(f"Jumlah barcode yang gagal diproses: {failed_processing_count}")

            final_message = f"Proses selesai!\n" \
                            f"Total berhasil: {successful_processing_count}\n" \
                            f"Baru digenerate: {generated_new_barcode_count}\n" \
                            f"Gagal: {failed_processing_count}"
            
            self.status_label.config(text="Status: Selesai!")
            self.results_label.config(text=final_message)
            messagebox.showinfo("Proses Selesai", "Pembuatan barcode berhasil!\nLihat log untuk detail.")

        except FileNotFoundError:
            self.log_message(f"Error: File input '{input_file}' tidak ditemukan.")
            messagebox.showerror("Error", f"File input '{input_file}' tidak ditemukan.")
        except pd.errors.EmptyDataError:
            self.log_message(f"Error: File Excel '{input_file}' kosong atau tidak memiliki data.")
            messagebox.showerror("Error", f"File Excel '{input_file}' kosong atau tidak memiliki data.")
        except KeyError as e:
            self.log_message(f"Error: Kolom yang dibutuhkan tidak ditemukan di Excel. Pastikan ada kolom '{BARCODE_COLUMN_NAME}' dan '{PRODUCT_NAME_COLUMN_NAME}'. Detail: {e}")
            messagebox.showerror("Error", f"Kolom yang dibutuhkan tidak ditemukan di Excel. Pastikan ada kolom '{BARCODE_COLUMN_NAME}' dan '{PRODUCT_NAME_COLUMN_NAME}'.")
        except Exception as e:
            self.log_message(f"\nTerjadi kesalahan utama selama proses: {str(e)}")
            messagebox.showerror("Error", f"Terjadi kesalahan: {str(e)}")
        finally:
            self.root.after(0, lambda: self._set_gui_processing_state(False))

    def export_empty_format_thread(self):
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile="format_barcode_kosong.xlsx", 
            title="Simpan Format Excel Kosong Sebagai"
        )

        if not file_path: 
            self.log_message("Ekspor format dibatalkan oleh pengguna.")
            return

        self._set_gui_processing_state(True) 
        self.log_message("Memulai proses ekspor format Excel kosong...")
        threading.Thread(target=self._export_empty_format_process, args=(file_path,)).start()

    def _export_empty_format_process(self, output_file_path): 
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Format Barcode Produk"
            
            headers = [PRODUCT_NAME_COLUMN_NAME, BARCODE_COLUMN_NAME]
            ws.append(headers)
            
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 25

            wb.save(output_file_path)
            self.log_message(f"Format Excel kosong berhasil diekspor ke: '{output_file_path}'")
            messagebox.showinfo("Export Berhasil", f"File format Excel kosong berhasil diekspor ke:\n{output_file_path}")
        except Exception as e:
            self.log_message(f"Error saat mengekspor format Excel kosong: {str(e)}")
            messagebox.showerror("Export Gagal", f"Gagal mengekspor format Excel kosong: {str(e)}")
        finally:
            self.root.after(0, lambda: self._set_gui_processing_state(False))


# --- Main Application ---
if __name__ == "__main__":
    root = tk.Tk()
    app = BarcodeApp(root)
    root.mainloop()
